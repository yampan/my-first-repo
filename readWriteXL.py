### readWriteXl.py
'''
セルの座標 (x, y) は 1 始まりで指定します。
openpyxl で読み書きできるのは .xlsx 形式のファイルのみです。
'''
# module
from openpyxl import load_workbook

def readXl(fn, deb=0):
    pass

# =======================================================
#  Excel ファイルの読み込み

def openXl():
    # Excel ファイルを開く
    workbook = load_workbook('your_excel_file.xlsx')

    # シートを取得
    sheet = workbook['Sheet1']  # シート名を指定

    # セル (x, y) の値を取得 (x, y は 1 始まり)
    x = 2  # 例: 2 行目
    y = 3  # 例: 3 列目
    cell_value = sheet.cell(row=x, column=y).value
    print(f"セル({x}, {y}) の値: {cell_value}")

    # すべてのセルを読み込む (ジェネレータ)
    for row in sheet.iter_rows():
        for cell in row:
            print(cell.value)

# ==============================================================
# Excel ファイルへの書き込み

def writeXl():
    # Excel ファイルを開く
    workbook = load_workbook('your_excel_file.xlsx')

    # シートを取得
    sheet = workbook['Sheet1']

    # セル (x, y) に値を書き込む (x, y は 1 始まり)
    x = 2  # 例: 2 行目
    y = 3  # 例: 3 列目
    sheet.cell(row=x, column=y).value = '新しい値'

    # Excel ファイルを保存
    workbook.save('your_excel_file.xlsx')

# ================================================================
# 特定のセル範囲の読み書き

def readWrite():

    # Excel ファイルを開く
    workbook = load_workbook('your_excel_file.xlsx')

    # シートを取得
    sheet = workbook['Sheet1']

    # セル範囲 (x1, y1) から (x2, y2) の値を読み込む
    x1 = 2
    y1 = 3
    x2 = 4
    y2 = 5
    for row in sheet.iter_rows(min_row=x1, min_col=y1, max_row=x2, max_col=y2):
        for cell in row:
            print(cell.value)

    # セル範囲 (x1, y1) から (x2, y2) に値を書き込む
    x1 = 2
    y1 = 3
    x2 = 4
    y2 = 5
    values = [['a', 'b', 'c'], ['d', 'e', 'f']]  # 書き込む値のリスト
    for i, row in enumerate(sheet.iter_rows(min_row=x1, min_col=y1, max_row=x2, max_col=y2)):
        for j, cell in enumerate(row):
            cell.value = values[i][j]

    # Excel ファイルを保存
    workbook.save('your_excel_file.xlsx')

# ================
# my_module.py


if __name__ == "__main__":
    print("このスクリプトは直接実行されました")
    